import openpyxl
import os

def extract_title_and_artist_from_excel(path, sheet_name):
    data_list = []
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheet_name]

    # Assuming column 1 is 'title' and column 3 is 'artist'
    for row in sheet.iter_rows(min_row=2, values_only=True):
        title, _, artist, *_ = row  # Extract title and artist from the row
        if title != None:
            data_list.append({'title': title, 'artist': artist})

    wb.close()
    return data_list

def extract_uri_from_excel(path, sheet_name):
    uri_list = []
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        spotify_uri = row[5]  # Extract title and artist from the row
        if spotify_uri and spotify_uri != 'NOT FOUND':
            uri_list.append(spotify_uri)

    wb.close()
    return uri_list

def write_spotify_uri_to_excel(uri_list, path, sheet_name):
    wb = openpyxl.load_workbook(path)                           
    sheet = wb[sheet_name]

    # Start writing from the second row (skipping the header row)
    for i, spotify_uri in enumerate(uri_list, start=2):
        sheet.cell(row=i, column=6, value=spotify_uri)

    wb.save(path)
    wb.close()


def load_or_create_output_sheet(output_file, output_sheet_name, header_row):
    # Check if output file exists
    if os.path.exists(output_file):
        output_workbook = openpyxl.load_workbook(output_file)
        if output_sheet_name in output_workbook.sheetnames:
            output_sheet = output_workbook[output_sheet_name]
        else:
            output_sheet = output_workbook.create_sheet(output_sheet_name)
    else:
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.create_sheet(output_sheet_name)

    # Copy the header row to the output sheet if it's empty
    print('max row is for output file is:', output_sheet.max_row)
    print('output file is', output_file)
    print(output_sheet.max_row)
    if output_sheet.max_row == 0:
        print('WE ARE APPENDING A NEW ROW')
        output_sheet.append(header_row)
        print(header_row)

    return output_workbook, output_sheet

def separate_valid_uris_and_invalid_uris(input_file, output_file, sheet_name):
    # Load the workbook and active sheet from the input file
    input_workbook = openpyxl.load_workbook(input_file)
    input_sheet = input_workbook[sheet_name]

    header_row = next(input_sheet.iter_rows(values_only=True))
    output_workbook, output_sheet = load_or_create_output_sheet(output_file, sheet_name, header_row)
    output_sheet.delete_rows(idx=2, amount=output_sheet.max_row - 1) #to ensure we are at a fresh start each time
    valid_rows = []

    # Copy only the rows with a value of 'NOT FOUND' in the Spotify uri column else append a valid row to 
    for row in input_sheet.iter_rows(min_row=2, values_only=True):
        spotify_uri = row[5]  # Assuming column 6 is the 'Spotify uri' column
        if spotify_uri == 'NOT FOUND':
            output_sheet.append(row)
        else:
            valid_rows.append(row)
    #Delete all rows except top row from original
    input_sheet.delete_rows(idx=2, amount=input_sheet.max_row - 1)
    #put back in valid rows
    for row in valid_rows:
        input_sheet.append(row)

    # Save workbooks to file path provided
    output_workbook.save(output_file)
    input_workbook.save(input_file)

def remove_parentheses_from_column(input_file, output_file, column_index, sheet_name):
    # Load the workbook and active sheet from the input file
    input_workbook = openpyxl.load_workbook(input_file)
    input_sheet = input_workbook[sheet_name]

    header_row = next(input_sheet.iter_rows(values_only=True))
    output_workbook, output_sheet = load_or_create_output_sheet(output_file, sheet_name, header_row)
    # Copy the data with parentheses removed to the output sheet
    for row in input_sheet.iter_rows(min_row=2, values_only=True):
        original_value = row[column_index - 1]
        if isinstance(original_value, str) and "(" in original_value:
            modified_value = original_value.split('(')[0].strip()
        else:
            modified_value = original_value
        row = list(row)
        row[column_index - 1] = modified_value
        output_sheet.append(row)

    # Save the new workbook to the output file
    output_workbook.save(output_file)

def get_sheet_names(path):
    workbook = openpyxl.load_workbook(path)
    return workbook.sheetnames